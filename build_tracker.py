#!/usr/bin/env python3
"""Build a side-by-side copy-recommendations tracker as .xlsx."""
import pathlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Copy Recommendations"

HEADERS = ["#", "Source pass", "Page", "Element", "Current copy / state",
           "Proposed edit / action", "Why (principle)", "Priority", "Status"]

# Status legend: OPEN / TEST / VERIFY / SHIPPED
rows = [
    # ---- Homepage, OPEN copy edits (Pass 3 / 06-29) ----
    ["P3 06-29", "Homepage", "Hero H1",
     "Get found by local customers, and chosen before your competitors.",
     "Get found by local customers — and chosen before the competitor down the street. Live in 24 hours.",
     "Specificity over vagueness: restore the timeframe the current H1 dropped.", "High", "OPEN"],
    ["P3 06-29", "Homepage", "Hero H1 — test variant A",
     "(same as above)",
     "Get found on Google and chosen by local customers — your site live in 24 hours.",
     "Outcome + timeframe; A/B against current.", "Test", "TEST"],
    ["P3 06-29", "Homepage", "Hero H1 — test variant B",
     "(same as above)",
     "The website an agency charges $5,000 for. Found-first, live in 24 hours, from $99/month.",
     "Price-contrast angle.", "Test", "TEST"],
    ["P3 06-29", "Homepage", "Hero H1 — test variant C",
     "(same as above)",
     "See your new website before you pay a cent — live in 24 hours.",
     "Leads with the free-preview risk-reversal hook.", "Test", "TEST"],
    ["P3 06-29", "Homepage", "Primary CTA button",
     "[CONFIRM actual button label — not readable from page]",
     "Get my free preview",
     "CTA = action verb + what they get; put the free-preview hook in the button.", "High", "OPEN"],
    ["P3 06-29", "Homepage", "About-Matt closing CTA",
     "If you want to see what I can do for your business, let's chat.",
     "See a free preview of your site — no commitment.",
     "\"Let's chat\" is a weak/vague CTA; end the strongest block with a concrete, zero-risk action.", "Med", "OPEN"],
    ["P3 06-29", "Homepage", "Eyebrow / bundle line",
     "Everything other agencies upsell you on - website, SEO, social, reviews - in one low monthly price.",
     "Everything agencies sell you separately — website, SEO, social, reviews — in one monthly price from $99.",
     "Specificity ($99 beats \"low\") + cleaner em-dashes.", "Med", "OPEN"],
    ["P3 06-29", "Homepage", "Duplicate tagline",
     "\"Agency results, a fraction of the price\" appears twice on the page.",
     "Keep once as a section divider; remove the duplicate.",
     "One idea per section; repetition dilutes.", "Low", "OPEN"],
    ["P3 06-29", "Homepage", "Frank testimonial — label",
     "Accordion player · Toronto",
     "Reconcile the location to match how Frank is referenced elsewhere (Niagara / his own site).",
     "Honest, consistent detail (testimonial wording itself is strong — leave it).", "Low", "OPEN"],

    # ---- Sub-pages, OPEN (Pass 3) ----
    ["P3 06-29", "Niagara + Affordable", "Proof / testimonial",
     "None — no testimonial, no founder block, no proof (homepage has all three).",
     "Add Frank's testimonial + a \"Built by Matt, here in Welland\" trust line with the Jetta Grove proof.",
     "These are your SEO landing pages; they convert worse than the homepage today.", "High", "OPEN"],
    ["P3 06-29", "Sub-pages (all)", "Guarantee placement",
     "Guarantee appears only in the bottom CTA.",
     "Move the guarantee line into the hero on each sub-page.",
     "Risk reversal belongs above the fold.", "Med", "OPEN"],
    ["P3 06-29", "One-Page page", "Hero H1",
     "[CONFIRM — may be \"When a one-page website works well.\"]",
     "One-page websites that make the next step obvious.",
     "Keyword + outcome-led H1 for \"one-page websites\" search intent.", "Med", "VERIFY"],

    # ---- CRO structural (Pass 3) ----
    ["P3 06-29 (CRO)", "Homepage", "CTA repetition",
     "Primary CTA not repeated down a now-long (~850-word) page.",
     "Repeat the primary CTA after the table, after the testimonial, after About, and after pricing.",
     "Long pages lose visitors between CTAs.", "High", "OPEN"],
    ["P3 06-29 (CRO)", "Homepage", "Booking option",
     "Form only — \"Tell me about your business and I'll build you a free preview\" (you reply later).",
     "Add an instant \"Book a 15-min call\" option (Cal.com) beside the form.",
     "High-intent visitors can act now instead of waiting; a competitor already books instantly.", "High", "OPEN"],
    ["P3 06-29 (CRO)", "Homepage", "FAQ section",
     "None.",
     "Add a short FAQ: what you need from me / is a 24-hour (AI) site any good / can I cancel / do I own it.",
     "Concentrates objection-handling right before the CTA.", "Med", "OPEN"],
    ["P3 06-29 (CRO)", "Homepage", "Review stars",
     "None.",
     "Add Google review stars + count near the hero once you have 5–10 (ask Frank first).",
     "Trust signal competitors lead with; you have none yet.", "Med", "OPEN"],
    ["P3 06-29 (CRO)", "Technical", "Meta title length",
     "Custom Websites in 24 Hours + Local Growth Marketing | Well and Good Websites (77 chars)",
     "Shorten to <65, e.g. \"Custom Websites in 24 Hours | Well & Good Websites\".",
     "Avoids truncation in Google results.", "Med", "OPEN"],
    ["P3 06-29 (CRO)", "Technical", "Image alt text",
     "Missing (flagged: no_image_alt).",
     "Add descriptive alt text to all images.",
     "Accessibility + local image SEO.", "Low", "OPEN"],
    ["P3 06-29 (CRO)", "Technical", "Analytics",
     "GA4 / Hotjar were failing to load in an earlier check.",
     "Verify tracking fires before running any A/B test.",
     "You can't measure a test if events don't fire.", "High", "OPEN"],

    # ---- Earlier passes, now SHIPPED (shown for progress) ----
    ["P1 06-27", "Homepage", "Hero (then)",
     "Custom sites in 24 hours. At 80% less than agency pricing.",
     "Move to an outcome-led hero.",
     "Headline should lead with the customer outcome, not a product attribute.", "—", "SHIPPED"],
    ["P1 06-27", "Homepage", "Guarantee",
     "None.",
     "\"Your site live in 24 hours, or your first month is free. Then I'll revise it until you're happy.\"",
     "Only risk-reversal in the competitive set.", "—", "SHIPPED"],
    ["P2 06-28", "Homepage", "Testimonial",
     "None (only a client's \"4.9-star\" reputation, not yours).",
     "Frank case study with a real, attributed result.",
     "Proof is the #1 conversion bottleneck.", "—", "SHIPPED"],
    ["P2 06-28", "Homepage", "About-Matt block",
     "Anonymous founder (\"I'll build you...\" with no name/face).",
     "First-person founder story + photo + mission.",
     "Solo + anonymous is a hard ask for a monthly commitment.", "—", "SHIPPED"],
    ["P2 06-28", "Homepage", "Founder proof",
     "Unattributed \"4.34M impressions / 163 keywords / Page 1.\"",
     "Attribute to your consultancy, Jetta Grove (founder track record), with a link.",
     "Unattributed bold stats read as less believable than none.", "—", "SHIPPED"],
    ["P2 06-28", "Homepage", "Eyebrow line",
     "Get more attention. Get more customers.",
     "Bundle line (everything agencies sell separately, one price).",
     "Replace filler only you-can't-say with a real differentiator.", "—", "SHIPPED"],
    ["P2 06-28", "Homepage", "github.io prototypes",
     "Two \"Recent Work\" demos on immeasurablematt.github.io read as fake client work.",
     "Relabel \"Concept build\" or move to real subdomains.",
     "Contradicted the \"agency results\" claim.", "—", "VERIFY"],
]

# ---- write ----
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="1A534F")
status_fills = {
    "OPEN": PatternFill("solid", fgColor="FFF2CC"),
    "TEST": PatternFill("solid", fgColor="DDEBF7"),
    "VERIFY": PatternFill("solid", fgColor="FCE4D6"),
    "SHIPPED": PatternFill("solid", fgColor="E2EFDA"),
}

ws.append(HEADERS)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = header_fill
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    c.border = border

for i, r in enumerate(rows, start=1):
    ws.append([i] + r)

widths = [4, 14, 18, 22, 46, 50, 40, 9, 10]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
    status = row[8].value
    if status in status_fills:
        row[8].fill = status_fills[status]
        row[8].font = Font(bold=True)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{ws.max_row}"

# legend sheet
leg = wb.create_sheet("How to use")
leg["A1"] = "How to use this tracker"; leg["A1"].font = Font(bold=True, size=14)
notes = [
    "",
    "One row per recommendation across all copy passes. Filter the Status column to focus.",
    "",
    "Status legend:",
    "  OPEN — proposed, not yet done. Your action.",
    "  TEST — an A/B variant to try, not a straight swap.",
    "  VERIFY — confirm current state on the live site (I couldn't read it).",
    "  SHIPPED — already live; shown so you can see progress.",
    "",
    "Columns E and F are the side-by-side: existing copy vs proposed edit.",
    "Column G is the principle (from the copywriting / CRO skills) behind the change.",
    "",
    "Source pass key:",
    "  P1 06-27 = Live Copy Audit",
    "  P2 06-28 = Competitive Action Plan + ready-to-ship copy",
    "  P3 06-29 = Copy Pass #2 (copywriting + CRO)",
    "",
    "Add a 'Notes' or 'Decision' column on the right as you review — yours to edit.",
]
for i, n in enumerate(notes, start=2):
    leg[f"A{i}"] = n
leg.column_dimensions["A"].width = 70

out = str(pathlib.Path(__file__).parent / "Well-and-Good-Websites-Copy-Recommendations-Tracker.xlsx")
wb.save(out)
print("Saved:", out, "| rows:", len(rows))
